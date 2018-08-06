import tensorflow as tf
import numpy as numpy
import scipy.io as scio
from openpyxl import load_workbook



def add1():
    sess = tf.Session()
    placeholder_list=[None for _ in range(9)]
    scalar_names = ["MLMP_Preference_0","MLMP_Preference_1","MLMP_Preference_2", \
                    "D-DASH-1_Preference_0","D-DASH-1_Preference_1","D-DASH-1_Preference_2", \
                    "D-DASH-2_Preference_0","D-DASH-2_Preference_1","D-DASH-2_Preference_2"]
    for x in range(9):
        placeholder_list[x] = tf.placeholder(dtype=tf.float32, shape=[],name = "{}".format(scalar_names[x]))
        tf.summary.scalar("{}".format(scalar_names[x]), placeholder_list[x])
        # tf.summary.scalar("Preference-{}".format(x/3), placeholder_list[x])
        
        # tf.summary.scalar("{}".format(scalar_names[x]), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        
        # tf.summary.scalar("MLMP_Preference_0".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("MLMP_Preference_1".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("MLMP_Preference_2".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("D-DASH-1_Preference_0".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("D-DASH-1_Preference_1".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("D-DASH-1_Preference_2".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("D-DASH-2_Preference_0".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("D-DASH-2_Preference_1".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
        # tf.summary.scalar("D-DASH-2_Preference_2".format(x), placeholder_list[x],collections=["Preference_{}".format(x/3)])
    tbdir = "./tensorboad"
    savename = "exp_1"
    tf_writer = tf.summary.FileWriter(tbdir + '/{}'.format(savename))
   

    # for x in range (len(placeholder_list)):
    #     cur_place = tf.placeholder(dtype=tf.float32, shape=[],name = "algo_{}task_{}".format(x))
    wb = load_workbook('./data.xlsx')
    print(wb.get_sheet_names())
    current_sheet = wb.get_sheet_by_name("sum")
    
    for row in range (2,1301):
        # for col in range (2,11):
        #     merge_op=tf.summary.merge_all()
        #     item = current_sheet.cell(row=row, column=col).value
        #     cur_placeholder = placeholder_list[col-2]
        #     # print ("item is {}\t{}".format(type(item),item))
        #     # print ("cur_placeholder is {}".format(cur_placeholder))
        #     # print ("merge_op is {}".format(merge_op))
        #     # print ("value is {}".format(sess.run(cur_placeholder,feed_dict={cur_placeholder:item})))
        #     summary = sess.run([merge_op],feed_dict={cur_placeholder:item})[0]
        #     tf_writer.add_summary(summary,row-1 )
        merge_op=tf.summary.merge_all()
        feed_dic = {}
        for col in range (2,11):
            # print ("items is {}".format(current_sheet.cell(row=row, column=col).value))
            feed_dic[placeholder_list[col-2]] = current_sheet.cell(row=row, column=col).value
        # summary = sess.run([merge_op],feed_dict={placeholder_list:items})[0]
        summary = sess.run([merge_op],feed_dict = feed_dic)[0]
        tf_writer.add_summary(summary,row-1 )
        tf_writer.flush()
        print ("flush row :{} is over".format(row))
    # for sheet_name in  wb.get_sheet_names():
    #     current_sheet = wb.get_sheet_by_name(sheet_name)
    #     print ("current_sheet is {} ".format(current_sheet.title))
    #     cur_algo = current_sheet.title.split("_task")[0]
    #     cur_task = current_sheet.title.splite("task_")[1]
        
    #     item = current_sheet.cell(row=1, column=2)
    #     print ("processing metric is {]".format(item.value))
    #     for row in range (2,1301):
    #         current_value = current_sheet.cell(row=row, column=2).value
    #         # b4_too = current_sheet.cell(row=4, column=2)

    
    
    # # sess = tf.Session()
    
    
    # self.summary_placeholder_list=[None for _ in range(self.num_subpolicies*2)]
def add2():
    sess = tf.Session()
    placeholder_list=[None for _ in range(3)]
    scalar_names = ["Preference_0","Preference_1","Preference_2"]
    for x in range(3):
        placeholder_list[x] = tf.placeholder(dtype=tf.float32, shape=[],name = "{}".format(scalar_names[x]))
        tf.summary.scalar("{}".format(scalar_names[x]), placeholder_list[x])
    tbdir = "./tensorboad"
    algo_list = ["MLMP","D-DASH-1","D-DASH-2"]
    tf_writer_list = []
    for algo_index, algo in enumerate(algo_list):
        tf_writer = tf.summary.FileWriter(tbdir + '/{}'.format(algo))
        tf_writer_list.append(tf_writer)
    wb = load_workbook('./data.xlsx')
    print(wb.get_sheet_names())
    current_sheet = wb.get_sheet_by_name("Sheet1")
    for row in range (2,1301):
        merge_op=tf.summary.merge_all()
        feed_dic = {}
        algo_point = 0
        for col in range (2,11):
            
            feed_dic[placeholder_list[(col-2)%3]] = current_sheet.cell(row=row, column=col).value
            if col %3 == 1:
                summary = sess.run([merge_op],feed_dict = feed_dic)[0]
                tf_writer_list[algo_point].add_summary(summary,row-1 )  
                tf_writer_list[algo_point].flush()
                algo_point+=1
        print ("flush row :{} is over".format(row))   


add2()
